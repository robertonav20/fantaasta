#!/usr/bin/env python3
"""Aggiorna i fogli quotazioni Fantacalcio di un workbook esistente.

Fogli modificati esclusivamente:
- Tutti
- Portieri
- Difensori
- Centrocampisti
- Attaccanti
- Ceduti

Il file sorgente viene scaricato dall'endpoint Fantacalcio usando un header Cookie
configurabile. Tutti gli altri fogli del workbook destinazione restano invariati.
"""

from __future__ import annotations

import argparse
import copy
import io
import json
import os
import re
import shutil
import sys
import tempfile
import unicodedata
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

SCRIPT_VERSION = "2026-08-19-catalog-timestamp-v8.3"

SHEETS_TO_UPDATE = (
    "Tutti",
    "Portieri",
    "Difensori",
    "Centrocampisti",
    "Attaccanti",
    "Ceduti",
)

EXPECTED_HEADERS = (
    "Id",
    "R",
    "RM",
    "Nome",
    "Squadra",
    "Qt.A",
    "Qt.I",
    "Diff.",
    "Qt.A M",
    "Qt.I M",
    "Diff.M",
    "FVM",
    "FVM M",
)

DATA_FIRST_COL = 1
DATA_LAST_COL = 13  # A:M
HEADER_ROW = 2


class UpdateError(RuntimeError):
    pass


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise UpdateError(f"File di configurazione non trovato: {path}")

    try:
        config = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise UpdateError(f"JSON non valido in {path}: {exc}") from exc

    required = ("source_url", "players_file")
    missing = [key for key in required if not config.get(key)]
    if missing:
        raise UpdateError("Parametri mancanti nel config: " + ", ".join(missing))

    return config


def resolve_path(value: str, config_dir: Path) -> Path:
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = config_dir / path
    return path.resolve()


def get_cookie(config: dict[str, Any]) -> str:
    # La variabile d'ambiente evita di salvare il cookie su disco.
    env_name = str(config.get("cookie_env", "FANTACALCIO_COOKIE"))
    cookie = os.environ.get(env_name) or str(config.get("cookie", ""))
    cookie = cookie.strip()
    if not cookie:
        raise UpdateError(
            f"Cookie mancante. Imposta '{env_name}' oppure il parametro 'cookie' nel config."
        )
    return cookie


def download_excel(config: dict[str, Any], url: str, label: str) -> bytes:
    cookie = get_cookie(config)
    timeout = int(config.get("timeout_seconds", 30))

    headers = {
        "Cookie": cookie,
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
        "User-Agent": str(
            config.get(
                "user_agent",
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/136.0 Safari/537.36",
            )
        ),
    }

    with requests.Session() as session:
        response = session.get(url, headers=headers, timeout=timeout, allow_redirects=True)

    if response.status_code in (401, 403):
        raise UpdateError(
            f"{label}: download non autorizzato (HTTP {response.status_code}). "
            "Verifica/aggiorna il cookie."
        )
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        raise UpdateError(f"{label}: download fallito: HTTP {response.status_code}") from exc

    content = response.content
    if len(content) < 4 or content[:2] != b"PK":
        content_type = response.headers.get("Content-Type", "sconosciuto")
        raise UpdateError(
            f"{label}: la risposta non sembra un file XLSX valido "
            f"(Content-Type: {content_type}, {len(content)} byte)."
        )

    return content


def download_text(
    config: dict[str, Any], url: str, label: str, use_cookie: bool = False
) -> str:
    """Scarica una pagina HTML opzionale usata per arricchire players.json."""
    timeout = int(config.get("timeout_seconds", 30))
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/json;q=0.9,*/*;q=0.8",
        "User-Agent": str(
            config.get(
                "user_agent",
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/136.0 Safari/537.36",
            )
        ),
    }
    if "understat.com" in url.casefold():
        headers.update(
            {
                "Accept-Language": "en-US,en;q=0.9,it;q=0.8",
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
                "Referer": "https://understat.com/",
            }
        )
    if use_cookie:
        headers["Cookie"] = get_cookie(config)

    with requests.Session() as session:
        response = session.get(url, headers=headers, timeout=timeout, allow_redirects=True)

    if response.status_code in (401, 403):
        raise UpdateError(f"{label}: accesso negato (HTTP {response.status_code}).")
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        raise UpdateError(f"{label}: download fallito: HTTP {response.status_code}") from exc

    response.encoding = response.encoding or response.apparent_encoding or "utf-8"
    text = response.text
    if len(text.strip()) < 100:
        raise UpdateError(f"{label}: risposta HTML troppo corta o vuota.")
    return text


def download_source(config: dict[str, Any]) -> bytes:
    return download_excel(config, str(config["source_url"]), "Quotazioni")


def sheet_lookup(workbook) -> dict[str, str]:
    return {name.casefold(): name for name in workbook.sheetnames}


def validate_source_workbook(source_wb) -> dict[str, str]:
    lookup = sheet_lookup(source_wb)
    resolved: dict[str, str] = {}

    for expected_name in SHEETS_TO_UPDATE:
        real_name = lookup.get(expected_name.casefold())
        if real_name is None:
            raise UpdateError(f"Foglio '{expected_name}' mancante nel file sorgente.")

        ws = source_wb[real_name]
        headers = tuple(ws.cell(HEADER_ROW, col).value for col in range(1, DATA_LAST_COL + 1))
        if headers != EXPECTED_HEADERS:
            raise UpdateError(
                f"Struttura inattesa nel foglio '{real_name}'. "
                f"Header atteso: {EXPECTED_HEADERS}; trovato: {headers}"
            )
        resolved[expected_name] = real_name

    return resolved



def count_data_rows(ws) -> int:
    return sum(
        1
        for row in range(3, ws.max_row + 1)
        if ws.cell(row=row, column=1).value is not None
    )


STATS_ID_ALIASES = {"id", "codice", "codice calciatore", "id calciatore"}
STATS_NAME_ALIASES = {"nome", "calciatore", "giocatore"}


def _norm_text(value: Any) -> str:
    return " ".join(str(value or "").strip().casefold().replace("_", " ").split())


def _player_id_key(value: Any) -> str:
    if value is None:
        return ""
    try:
        number = float(value)
        if number.is_integer():
            return str(int(number))
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _find_stats_header_row(ws) -> tuple[int, dict[str, int]] | None:
    for row in range(1, min(ws.max_row, 25) + 1):
        values = {
            _norm_text(ws.cell(row, col).value): col
            for col in range(1, ws.max_column + 1)
            if ws.cell(row, col).value not in (None, "")
        }
        id_col = next((values[x] for x in STATS_ID_ALIASES if x in values), None)
        name_col = next((values[x] for x in STATS_NAME_ALIASES if x in values), None)
        if id_col and name_col:
            headers: dict[str, int] = {}
            for col in range(1, ws.max_column + 1):
                raw = ws.cell(row, col).value
                if raw in (None, ""):
                    continue
                label = str(raw).strip()
                if label in headers:
                    label = f"{label} ({col})"
                headers[label] = col
            return row, headers
    return None


def parse_stats_workbook(source_bytes: bytes, season: str) -> dict[str, dict[str, Any]]:
    try:
        wb = load_workbook(io.BytesIO(source_bytes), read_only=False, data_only=True)
    except Exception as exc:
        raise UpdateError(f"Statistiche {season}: XLSX non leggibile: {exc}") from exc

    by_id: dict[str, dict[str, Any]] = {}
    try:
        for ws in wb.worksheets:
            found = _find_stats_header_row(ws)
            if not found:
                continue
            header_row, headers = found
            normalized = {_norm_text(label): col for label, col in headers.items()}
            id_col = next((normalized[x] for x in STATS_ID_ALIASES if x in normalized), None)
            name_col = next((normalized[x] for x in STATS_NAME_ALIASES if x in normalized), None)
            if not id_col or not name_col:
                continue

            for row in range(header_row + 1, ws.max_row + 1):
                player_id = ws.cell(row, id_col).value
                name = ws.cell(row, name_col).value
                key = _player_id_key(player_id)
                if not key or not name:
                    continue
                details = {
                    label: _json_value(ws.cell(row, col).value)
                    for label, col in headers.items()
                }
                # In caso di piu' fogli, il record piu' ricco vince.
                current = by_id.get(key)
                if current is None or sum(v not in (None, "") for v in details.values()) > sum(
                    v not in (None, "") for v in current.values()
                ):
                    by_id[key] = details
    finally:
        wb.close()

    if not by_id:
        raise UpdateError(
            f"Statistiche {season}: nessun record trovato. "
            "Formato del file inatteso oppure statistiche non ancora disponibili."
        )
    return by_id


def load_existing_stats(json_file: Path) -> dict[str, dict[str, dict[str, Any]]]:
    result: dict[str, dict[str, dict[str, Any]]] = {}
    if not json_file.exists():
        return result
    try:
        payload = json.loads(json_file.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return result
    for role in ROLE_SHEETS:
        for player in payload.get(role, []) if isinstance(payload, dict) else []:
            key = _player_id_key(player.get("id"))
            stats = player.get("stats")
            if key and isinstance(stats, dict):
                result[key] = copy.deepcopy(stats)
    return result


def download_stats_catalogs(
    config: dict[str, Any], existing_stats: dict[str, dict[str, dict[str, Any]]]
) -> tuple[dict[str, dict[str, dict[str, Any]]], list[str]]:
    sources = config.get("stats_sources") or []
    if not isinstance(sources, list):
        raise UpdateError("'stats_sources' deve essere una lista nel file di configurazione.")

    merged = copy.deepcopy(existing_stats)
    loaded: list[str] = []
    for item in sources:
        if not isinstance(item, dict) or not item.get("season") or not item.get("url"):
            raise UpdateError("Ogni elemento di 'stats_sources' richiede 'season' e 'url'.")
        season = str(item["season"]).strip()
        required = bool(item.get("required", False))
        try:
            content = download_excel(config, str(item["url"]), f"Statistiche {season}")
            stats = parse_stats_workbook(content, season)
        except (UpdateError, requests.RequestException) as exc:
            if required:
                raise
            print(f"ATTENZIONE: {exc}. Mantengo eventuali dati {season} gia' presenti.", file=sys.stderr)
            continue
        for player_id, details in stats.items():
            merged.setdefault(player_id, {})[season] = details
        loaded.append(season)
        print(f"Statistiche {season}: {len(stats)} giocatori")
    return merged, loaded


ROLE_SHEETS = {
    "P": "Portieri",
    "D": "Difensori",
    "C": "Centrocampisti",
    "A": "Attaccanti",
}


def _json_number(value: Any) -> int | float:
    number = float(value or 0)
    if number.is_integer():
        return int(number)
    return number


def _json_value(value: Any) -> Any:
    """Converte un valore Excel in un tipo JSON serializzabile senza perdere i numeri."""
    if value is None or isinstance(value, (str, bool)):
        return value
    if isinstance(value, (int, float)):
        return _json_number(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)




def _name_tokens(value: Any) -> list[str]:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.casefold().replace("’", "'")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return [token for token in text.split() if token]


def _team_key(value: Any) -> str:
    tokens = [t for t in _name_tokens(value) if t not in {"fc", "ac", "calcio", "1909", "1913"}]
    text = " ".join(tokens)
    aliases = {
        "internazionale": "inter",
        "internazionale milano": "inter",
        "inter milan": "inter",
        "hellas verona": "verona",
        "ssc napoli": "napoli",
        "ss lazio": "lazio",
        "as roma": "roma",
        "juventus turin": "juventus",
    }
    return aliases.get(text, text)


def _fc_name_signature(value: Any) -> tuple[list[str], str]:
    tokens = _name_tokens(value)
    hint = ""
    if len(tokens) > 1 and len(tokens[-1]) <= 2:
        hint = tokens[-1]
        tokens = tokens[:-1]
    return tokens, hint


def _names_compatible(fc_name: str, external_name: str) -> bool:
    fc_tokens, hint = _fc_name_signature(fc_name)
    ext_tokens = _name_tokens(external_name)
    if not fc_tokens or not ext_tokens:
        return False
    if fc_tokens == ext_tokens:
        return True
    if len(fc_tokens) <= len(ext_tokens) and ext_tokens[-len(fc_tokens):] == fc_tokens:
        if not hint:
            return True
        given = ext_tokens[:-len(fc_tokens)]
        return any(token.startswith(hint) for token in given)
    return False


def load_player_identities(workbook_file: Path) -> list[dict[str, Any]]:
    """Identita' minime usate per il matching conservativo delle fonti esterne."""
    wb = load_workbook(workbook_file, read_only=False, data_only=True)
    lookup = sheet_lookup(wb)
    players: list[dict[str, Any]] = []
    try:
        for role, expected_sheet in ROLE_SHEETS.items():
            real_name = lookup.get(expected_sheet.casefold())
            if not real_name:
                continue
            ws = wb[real_name]
            headers = {
                str(ws.cell(HEADER_ROW, col).value): col
                for col in range(DATA_FIRST_COL, DATA_LAST_COL + 1)
                if ws.cell(HEADER_ROW, col).value is not None
            }
            if not all(k in headers for k in ("Id", "Nome", "Squadra")):
                continue
            for row in range(HEADER_ROW + 1, ws.max_row + 1):
                player_id = ws.cell(row, headers["Id"]).value
                name = ws.cell(row, headers["Nome"]).value
                if player_id is None or not name:
                    continue
                players.append(
                    {
                        "id": _player_id_key(player_id),
                        "name": str(name).strip(),
                        "team": str(ws.cell(row, headers["Squadra"]).value or "").strip(),
                        "role": role,
                    }
                )
    finally:
        wb.close()
    return players


def match_player_id(
    external_name: str,
    identities: list[dict[str, Any]],
    external_team: str | None = None,
) -> str | None:
    """Restituisce un Id solo se il match e' univoco; altrimenti ignora il record."""
    exact = [p for p in identities if _name_tokens(p["name"]) == _name_tokens(external_name)]
    candidates = exact or [p for p in identities if _names_compatible(p["name"], external_name)]
    if not candidates:
        return None

    if external_team:
        team_key = _team_key(external_team)
        same_team = [p for p in candidates if _team_key(p["team"]) == team_key]
        if len(same_team) == 1:
            return same_team[0]["id"]
        if not same_team:
            return None
        candidates = same_team

    unique_ids = {p["id"] for p in candidates}
    return next(iter(unique_ids)) if len(unique_ids) == 1 else None


def _number_or_text(value: Any) -> Any:
    if value in (None, ""):
        return value
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    return int(number) if number.is_integer() else round(number, 4)



def _scan_js_string_literal(text: str, start: int) -> tuple[str, int]:
    """Restituisce una stringa JS quotata completa e l'indice successivo."""
    if start >= len(text) or text[start] not in ("'", '"'):
        raise UpdateError("Understat: argomento JSON.parse non e' una stringa JS.")
    quote = text[start]
    escaped = False
    i = start + 1
    while i < len(text):
        ch = text[i]
        if escaped:
            escaped = False
        elif ch == "\\":
            escaped = True
        elif ch == quote:
            return text[start : i + 1], i + 1
        i += 1
    raise UpdateError("Understat: stringa JS playersData non terminata.")


def _scan_balanced_json(text: str, start: int) -> str:
    """Estrae un array/oggetto JSON bilanciando parentesi e stringhe."""
    if start >= len(text) or text[start] not in "[{":
        raise UpdateError("Understat: payload JSON diretto non valido.")
    opening = text[start]
    closing = "]" if opening == "[" else "}"
    depth = 0
    in_string = False
    escaped = False
    quote = ""
    for i in range(start, len(text)):
        ch = text[i]
        if in_string:
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == quote:
                in_string = False
            continue
        if ch in ("'", '"'):
            in_string = True
            quote = ch
        elif ch == opening:
            depth += 1
        elif ch == closing:
            depth -= 1
            if depth == 0:
                return text[start : i + 1]
    raise UpdateError("Understat: payload JSON playersData non terminato.")


def _decode_understat_js_literal(literal: str) -> Any:
    """Decodifica la stringa JS usata da Understat dentro JSON.parse()."""
    import ast

    try:
        # Understat usa storicamente escape JS/Python compatibili, inclusi \xNN.
        decoded = ast.literal_eval(literal)
    except (SyntaxError, ValueError) as exc:
        raise UpdateError(f"Understat: stringa playersData non decodificabile: {exc}") from exc
    if not isinstance(decoded, str):
        raise UpdateError("Understat: playersData decodificato non e' testo JSON.")
    try:
        return json.loads(decoded)
    except json.JSONDecodeError as exc:
        raise UpdateError(f"Understat: JSON playersData non valido: {exc}") from exc


def _understat_is_challenge(html: str) -> bool:
    # Evita il falso positivo su semplici riferimenti a Cloudflare presenti
    # anche in pagine normali (analytics/CDN). Cerchiamo marker specifici.
    lowered = html.casefold()
    challenge_markers = (
        "just a moment",
        "cf-chl-",
        "challenge-platform",
        "verify you are human",
        "verification successful",
        "turnstile",
        "captcha",
        "access denied",
        "enable javascript and cookies",
    )
    return any(marker in lowered for marker in challenge_markers)


def _understat_has_payload(html: str, expected: tuple[str, ...] = ("playersData", "teamsData")) -> bool:
    return any(re.search(rf"\b{re.escape(name)}\b", html) for name in expected)




def _understat_find_player_rows_in_json(payload: Any) -> list[dict[str, Any]]:
    """Trova ricorsivamente il blocco giocatori in un payload JSON Understat."""
    best: list[dict[str, Any]] = []

    def walk(value: Any) -> None:
        nonlocal best
        if isinstance(value, list):
            dict_rows = [item for item in value if isinstance(item, dict)]
            if dict_rows:
                score = 0
                for row in dict_rows[:10]:
                    keys = set(row)
                    if "player_name" in keys or "player" in keys or "name" in keys:
                        score += 2
                    if keys & {"xG", "xA", "npxG", "xGChain", "xGBuildup", "games", "time"}:
                        score += 2
                if score >= 4 and len(dict_rows) > len(best):
                    best = dict_rows
            for item in value:
                walk(item)
        elif isinstance(value, dict):
            for item in value.values():
                walk(item)

    walk(payload)
    normalized: list[dict[str, Any]] = []
    for row in best:
        item = dict(row)
        if "player_name" not in item:
            candidate = item.get("player") or item.get("name")
            if isinstance(candidate, str):
                item["player_name"] = candidate
            elif isinstance(candidate, dict):
                item["player_name"] = candidate.get("name") or candidate.get("title") or ""
        normalized.append(item)
    return normalized


def _understat_rows_from_rendered_html(html: str) -> list[dict[str, Any]]:
    """Fallback DOM: legge una tabella giocatori gia' renderizzata da Chromium."""
    soup = BeautifulSoup(html, "html.parser")
    best: list[dict[str, Any]] = []

    aliases = {
        "apps": "games", "app": "games", "games": "games", "matches": "games",
        "min": "time", "mins": "time", "minutes": "time",
        "g": "goals", "goals": "goals",
        "npg": "npg", "a": "assists", "assists": "assists",
        "xg": "xG", "npxg": "npxG", "xa": "xA",
        "shots": "shots", "sh": "shots",
        "kp": "key_passes", "keypasses": "key_passes", "key_passes": "key_passes",
        "yc": "yellow_cards", "yellowcards": "yellow_cards",
        "rc": "red_cards", "redcards": "red_cards",
        "xgchain": "xGChain", "xgbuildup": "xGBuildup",
        "position": "position", "pos": "position",
    }

    for table in soup.find_all("table"):
        headers = []
        header_row = table.find("tr")
        if header_row:
            headers = [re.sub(r"[^a-z0-9_]+", "", cell.get_text(" ", strip=True).casefold())
                       for cell in header_row.find_all(["th", "td"])]
        rows: list[dict[str, Any]] = []
        for tr in table.find_all("tr"):
            player_link = tr.find("a", href=re.compile(r"/player/\d+"))
            if not player_link:
                continue
            cells = tr.find_all(["td", "th"])
            row: dict[str, Any] = {"player_name": player_link.get_text(" ", strip=True)}
            m = re.search(r"/player/(\d+)", player_link.get("href", ""))
            if m:
                row["id"] = m.group(1)
            team_link = tr.find("a", href=re.compile(r"/team/"))
            if team_link:
                row["team_title"] = team_link.get_text(" ", strip=True)
            for idx, cell in enumerate(cells):
                if idx >= len(headers):
                    continue
                key = aliases.get(headers[idx])
                if not key:
                    continue
                raw = cell.get("data-sort-value") or cell.get_text(" ", strip=True)
                raw = str(raw).replace("%", "").strip()
                if raw:
                    row[key] = raw
            rows.append(row)
        if len(rows) > len(best):
            best = rows
    return best


def _understat_inject_players(html: str, rows: list[dict[str, Any]]) -> str:
    if not rows:
        return html
    payload = json.dumps(rows, ensure_ascii=False, separators=(",", ":"))
    return html + f"\n<script>var playersData = {payload};</script>\n"


def _understat_validate_title_season(html: str, season: str) -> None:
    """Evita di salvare 2025/26 sotto 2026/27 quando Understat fa fallback di stagione."""
    m = re.search(r"<title[^>]*>(.*?)</title>", html, re.I | re.S)
    if not m:
        return
    title = re.sub(r"\s+", " ", m.group(1)).strip()
    found = re.search(r"(20\d{2})/(20\d{2})", title)
    expected = re.match(r"(20\d{2})/(\d{2})", season)
    if not found or not expected:
        return
    exp_full = f"{expected.group(1)}/20{expected.group(2)}"
    got = f"{found.group(1)}/{found.group(2)}"
    if got != exp_full:
        raise UpdateError(f"Understat {season}: la pagina restituisce la stagione {got}, quindi la stagione richiesta non e' ancora disponibile.")

def download_understat_browser(
    config: dict[str, Any],
    url: str,
    label: str,
    expected: tuple[str, ...] = ("playersData", "teamsData"),
) -> str:
    """Fallback browser reale: esegue JavaScript e restituisce l'HTML renderizzato."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise UpdateError(
            f"{label}: Playwright non installato. Aggiungi 'playwright' alle dipendenze "
            "e installa Chromium nel workflow."
        ) from exc

    timeout_ms = int(config.get("understat_browser_timeout_seconds", 35)) * 1000
    settle_ms = int(config.get("understat_browser_settle_ms", 2500))

    # GitHub Actions puo' avere Playwright installato ma non il browser scaricato.
    # Prima prova un executable esplicito/configurato o un Chrome/Chromium di sistema;
    # se non esiste, lascia che Playwright usi il proprio browser installato.
    configured_executable = str(
        config.get("understat_browser_executable", "")
        or os.environ.get("UNDERSTAT_BROWSER_EXECUTABLE", "")
    ).strip()
    browser_executable = configured_executable or next(
        (
            candidate
            for name in ("google-chrome", "google-chrome-stable", "chromium", "chromium-browser")
            if (candidate := shutil.which(name))
        ),
        "",
    )

    try:
        with sync_playwright() as pw:
            launch_kwargs: dict[str, Any] = {"headless": True}
            if browser_executable:
                launch_kwargs["executable_path"] = browser_executable
                print(f"{label}: uso browser di sistema {browser_executable}")
            else:
                print(f"{label}: uso Chromium gestito da Playwright")
            browser = pw.chromium.launch(**launch_kwargs)
            browser_ua = str(config.get("understat_browser_user_agent", "")).strip()
            if not browser_ua:
                browser_ua = (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    f"Chrome/{browser.version} Safari/537.36"
                )
            context = browser.new_context(
                user_agent=browser_ua,
                locale="en-US",
                timezone_id="Europe/Rome",
                viewport={"width": 1365, "height": 900},
                extra_http_headers={
                    "Accept-Language": "en-US,en;q=0.9,it;q=0.8",
                    "Referer": "https://understat.com/",
                },
            )
            page = context.new_page()
            network_json: list[Any] = []

            def _capture_response(resp: Any) -> None:
                try:
                    content_type = str(resp.headers.get("content-type", "")).casefold()
                    if "json" in content_type:
                        network_json.append(resp.json())
                except Exception:
                    pass

            page.on("response", _capture_response)
            response = page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            initial_status = response.status if response is not None else None

            # Anche una challenge HTTP 403 puo' risolversi via JavaScript.
            deadline_ms = min(timeout_ms, 15000)
            elapsed = 0
            html = page.content()
            while elapsed < deadline_ms and not _understat_has_payload(html, expected):
                page.wait_for_timeout(1000)
                elapsed += 1000
                html = page.content()
                if not _understat_is_challenge(html) and elapsed >= settle_ms:
                    # Pagina normale ma senza il payload atteso: inutile aspettare oltre.
                    break

            # Nuovo Understat: i vecchi globals playersData/teamsData possono non
            # esistere piu'. Recupera i dati dalle risposte XHR/fetch oppure dalla
            # tabella DOM renderizzata e reinseriscili nel formato atteso dal parser.
            if not _understat_has_payload(html, expected):
                candidate_rows: list[dict[str, Any]] = []
                for payload in network_json:
                    rows = _understat_find_player_rows_in_json(payload)
                    if len(rows) > len(candidate_rows):
                        candidate_rows = rows
                if not candidate_rows:
                    candidate_rows = _understat_rows_from_rendered_html(html)
                if candidate_rows:
                    print(f"{label}: recuperati {len(candidate_rows)} giocatori da XHR/DOM renderizzato")
                    html = _understat_inject_players(html, candidate_rows)

            context.close()
            browser.close()
    except UpdateError:
        raise
    except Exception as exc:
        raise UpdateError(f"{label}: fallback Chromium fallito: {exc}") from exc

    if len(html.strip()) < 100:
        raise UpdateError(f"{label}: HTML browser troppo corto o vuoto.")
    if _understat_is_challenge(html) and not _understat_has_payload(html, expected):
        raise UpdateError(
            f"{label}: challenge anti-bot ancora presente anche dopo Chromium."
        )
    if not _understat_has_payload(html, expected):
        title = ""
        match = re.search(r"<title[^>]*>(.*?)</title>", html, re.I | re.S)
        if match:
            title = re.sub(r"\s+", " ", match.group(1)).strip()
        status_detail = f", HTTP iniziale={initial_status}" if initial_status is not None else ""
        raise UpdateError(
            f"{label}: Chromium ha caricato la pagina ma non contiene "
            f"{', '.join(expected)} (title={title!r}, {len(html)} caratteri{status_detail})."
        )
    print(f"{label}: HTML recuperato con Chromium")
    return html


def download_understat_html(
    config: dict[str, Any],
    url: str,
    label: str,
    expected: tuple[str, ...] = ("playersData", "teamsData"),
) -> str:
    """Prima requests; se manca il payload/challenge, usa Chromium."""
    request_error: Exception | None = None
    try:
        html = download_text(config, url, label)
        if _understat_has_payload(html, expected) and not _understat_is_challenge(html):
            return html
        reason = (
            "challenge rilevata" if _understat_is_challenge(html)
            else f"payload {', '.join(expected)} assente"
        )
        print(f"ATTENZIONE: {label}: {reason}. Provo Chromium.", file=sys.stderr)
    except (UpdateError, requests.RequestException) as exc:
        request_error = exc
        print(f"ATTENZIONE: {label}: requests fallito ({exc}). Provo Chromium.", file=sys.stderr)

    if not bool(config.get("understat_browser_fallback", True)):
        if request_error:
            raise UpdateError(str(request_error))
        raise UpdateError(f"{label}: payload Understat assente e fallback Chromium disabilitato.")
    return download_understat_browser(config, url, label, expected)


def _understat_missing_payload_error(html: str) -> UpdateError:
    if _understat_is_challenge(html):
        return UpdateError(
            "Understat: ricevuta pagina anti-bot/challenge invece dei dati playersData."
        )
    title = ""
    match = re.search(r"<title[^>]*>(.*?)</title>", html, re.I | re.S)
    if match:
        title = re.sub(r"\s+", " ", match.group(1)).strip()
    detail = f", title={title!r}" if title else ""
    return UpdateError(
        f"Understat: playersData non trovato nel sorgente HTML ({len(html)} caratteri{detail})."
    )


def _parse_understat_variable(html: str, variable_name: str) -> Any:
    """Estrae una variabile JS Understat dal sorgente HTML grezzo."""
    marker_re = re.compile(
        rf"\b(?:var\s+|let\s+|const\s+)?{re.escape(variable_name)}\b"
    )
    markers = list(marker_re.finditer(html))
    if not markers:
        if variable_name == "playersData":
            raise _understat_missing_payload_error(html)
        raise UpdateError(f"Understat: {variable_name} non trovato nel sorgente HTML.")

    last_error: UpdateError | None = None
    for marker in markers:
        assignment = re.match(r"\s*=\s*", html[marker.end() :])
        if not assignment:
            continue
        pos = marker.end() + assignment.end()

        parse_match = re.match(r"JSON\.parse\s*\(\s*", html[pos:])
        if parse_match:
            literal_pos = pos + parse_match.end()
            if literal_pos < len(html) and html[literal_pos] in ("'", '"'):
                try:
                    literal, _ = _scan_js_string_literal(html, literal_pos)
                    return _decode_understat_js_literal(literal)
                except UpdateError as exc:
                    last_error = exc
                    continue

        if pos < len(html) and html[pos] in "[{":
            try:
                payload = _scan_balanced_json(html, pos)
                return json.loads(payload)
            except (UpdateError, json.JSONDecodeError) as exc:
                last_error = UpdateError(
                    f"Understat: JSON diretto {variable_name} non valido: {exc}"
                )
                continue

    if last_error:
        raise last_error
    raise UpdateError(f"Understat: {variable_name} trovato, ma formato non riconosciuto.")


def parse_understat_players(html: str) -> list[dict[str, Any]]:
    """Estrae playersData dal sorgente HTML pubblico di Understat."""
    data = _parse_understat_variable(html, "playersData")
    if not isinstance(data, list):
        raise UpdateError("Understat: playersData non e' una lista.")
    if not data:
        raise UpdateError("Understat: playersData presente ma vuoto.")
    return [row for row in data if isinstance(row, dict)]


def parse_understat_team_titles(html: str) -> list[str]:
    """Estrae i nomi squadra da teamsData per il fallback sulle pagine team."""
    try:
        data = _parse_understat_variable(html, "teamsData")
    except UpdateError:
        data = None

    titles: list[str] = []
    values: list[Any]
    if isinstance(data, dict):
        values = list(data.values())
    elif isinstance(data, list):
        values = data
    else:
        values = []

    for item in values:
        if not isinstance(item, dict):
            continue
        title = str(item.get("title") or item.get("team_title") or "").strip()
        if title and title not in titles:
            titles.append(title)

    # Secondo fallback: link team presenti direttamente nell'HTML.
    if not titles:
        for slug in re.findall(r"href=[\"'](?:https?://(?:www\.)?understat\.com)?/team/([^/\"'?#]+)", html, re.I):
            title = slug.replace("_", " ").strip()
            if title and title not in titles:
                titles.append(title)
    return titles


UNDERSTAT_FIELDS = (
    "games",
    "time",
    "goals",
    "xG",
    "assists",
    "xA",
    "shots",
    "key_passes",
    "yellow_cards",
    "red_cards",
    "position",
    "team_title",
    "npg",
    "npxG",
    "xGChain",
    "xGBuildup",
)


def understat_catalog_from_rows(
    rows: list[dict[str, Any]], season: str, identities: list[dict[str, Any]]
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        name = str(row.get("player_name") or "").strip()
        team = row.get("team_title")
        if isinstance(team, list):
            team = team[-1] if team else ""

        # Prima prova nome+squadra. Per lo storico consenti poi il solo nome,
        # ma esclusivamente se il match resta univoco: copre i trasferimenti.
        player_id = match_player_id(name, identities, str(team or ""))
        if not player_id:
            player_id = match_player_id(name, identities, None)
        if not player_id:
            continue

        details = {
            field: _number_or_text(row.get(field))
            for field in UNDERSTAT_FIELDS
            if field in row
        }
        minutes = float(details.get("time") or 0)
        if minutes > 0:
            xg = float(details.get("xG") or 0)
            xa = float(details.get("xA") or 0)
            details["xG90"] = round(xg * 90 / minutes, 3)
            details["xA90"] = round(xa * 90 / minutes, 3)
            details["xG+xA90"] = round((xg + xa) * 90 / minutes, 3)
        details["provider"] = "Understat"
        details["season"] = season
        result[player_id] = details
    if not result:
        raise UpdateError(f"Understat {season}: nessun match univoco col listone Fantacalcio.")
    return result


def understat_catalog(
    html: str, season: str, identities: list[dict[str, Any]]
) -> dict[str, dict[str, Any]]:
    return understat_catalog_from_rows(parse_understat_players(html), season, identities)


def _understat_year(source_url: str, season: str) -> str:
    match = re.search(r"/(20\d{2})(?:[/?#]|$)", source_url)
    if match:
        return match.group(1)
    match = re.match(r"(20\d{2})", season)
    if match:
        return match.group(1)
    raise UpdateError(f"Understat {season}: anno stagione non determinabile.")


def _understat_team_slug(title: str) -> str:
    # Understat usa normalmente spazi sostituiti da underscore nei path team.
    return requests.utils.quote(title.strip().replace(" ", "_"), safe="_-.")


def understat_team_fallback_catalog(
    config: dict[str, Any],
    source_url: str,
    league_html: str,
    season: str,
    identities: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """Fallback HTML: aggrega playersData dalle pagine delle singole squadre."""
    titles = parse_understat_team_titles(league_html)
    if not titles:
        raise UpdateError(
            f"Understat {season}: playersData lega assente e nessuna squadra ricavabile da teamsData."
        )

    year = _understat_year(source_url, season)
    base = "https://understat.com"
    rows: list[dict[str, Any]] = []
    successes = 0
    for title in titles[:30]:
        slug = _understat_team_slug(title)
        team_url = f"{base}/team/{slug}/{year}"
        try:
            team_html = download_understat_html(
                config,
                team_url,
                f"Understat {season} {title}",
                expected=("playersData",),
            )
            team_rows = parse_understat_players(team_html)
        except (UpdateError, requests.RequestException) as exc:
            print(f"ATTENZIONE: fallback Understat {title}: {exc}", file=sys.stderr)
            continue
        rows.extend(team_rows)
        successes += 1

    if not rows:
        raise UpdateError(
            f"Understat {season}: fallback pagine squadra fallito ({successes}/{len(titles)} squadre)."
        )
    print(
        f"Understat {season}: fallback HTML squadre usato ({successes}/{len(titles)} squadre, {len(rows)} righe)"
    )
    return understat_catalog_from_rows(rows, season, identities)



def _understat_api_url(source_url: str, season: str) -> str:
    """Costruisce l'endpoint JSON usato dal sito Understat corrente."""
    match = re.search(r"/league/([^/]+)/([0-9]{4})(?:[/?#]|$)", source_url)
    if match:
        league_slug, year = match.group(1), match.group(2)
    else:
        league_slug = "Serie_A"
        year_match = re.match(r"(20\d{2})", season)
        if not year_match:
            raise UpdateError(f"Understat {season}: impossibile determinare l'anno stagione.")
        year = year_match.group(1)
    return f"https://understat.com/getLeagueData/{league_slug}/{year}"


def download_understat_league_json(
    config: dict[str, Any], source_url: str, season: str, label: str
) -> dict[str, Any]:
    """Scarica il JSON usato dal frontend Understat corrente.

    Understat non incorpora piu' stabilmente playersData/teamsData nella pagina HTML.
    Il frontend corrente usa /getLeagueData/{league}/{season}; prima inizializziamo
    i cookie sulla home e poi usiamo l'header XMLHttpRequest richiesto dal sito.
    """
    timeout = int(config.get("timeout_seconds", 30))
    api_url = _understat_api_url(source_url, season)
    user_agent = str(
        config.get(
            "user_agent",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/136.0 Safari/537.36",
        )
    )
    base_headers = {
        "User-Agent": user_agent,
        "Accept-Language": "en-US,en;q=0.9,it;q=0.8",
        "Referer": source_url,
    }
    api_headers = {
        **base_headers,
        "Accept": "application/json,text/plain,*/*",
        "X-Requested-With": "XMLHttpRequest",
    }

    try:
        with requests.Session() as session:
            # Understat inizializza i cookie con una visita alla home.
            session.get(
                "https://understat.com/",
                headers={**base_headers, "Accept": "text/html,application/xhtml+xml,*/*;q=0.8"},
                timeout=timeout,
                allow_redirects=True,
            )
            response = session.get(
                api_url,
                headers=api_headers,
                timeout=timeout,
                allow_redirects=True,
            )
    except requests.RequestException as exc:
        raise UpdateError(f"{label}: endpoint JSON Understat non raggiungibile: {exc}") from exc

    if response.status_code in (401, 403):
        raise UpdateError(f"{label}: endpoint JSON Understat accesso negato (HTTP {response.status_code}).")
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        raise UpdateError(
            f"{label}: endpoint JSON Understat HTTP {response.status_code}."
        ) from exc

    try:
        data = response.json()
    except ValueError as exc:
        sample = response.text[:120].replace("\n", " ")
        raise UpdateError(
            f"{label}: endpoint JSON Understat ha restituito contenuto non JSON: {sample!r}"
        ) from exc

    if not isinstance(data, dict):
        raise UpdateError(f"{label}: risposta JSON Understat non e' un oggetto.")
    players = data.get("players")
    if not isinstance(players, list) or not players:
        raise UpdateError(f"{label}: risposta JSON Understat senza lista players.")

    print(f"{label}: endpoint JSON corrente caricato ({len(players)} giocatori)")
    return data


def understat_catalog_current(
    config: dict[str, Any],
    source_url: str,
    season: str,
    identities: list[dict[str, Any]],
    label: str,
) -> dict[str, dict[str, Any]]:
    """Parser Understat corrente con compatibilita' legacy HTML."""
    # 1) HTML: serve a validare che Understat stia davvero mostrando la stagione richiesta.
    html = download_text(config, source_url, label)
    _understat_validate_title_season(html, season)

    # 2) Compatibilita' con il vecchio formato HTML, se ancora presente.
    if _understat_has_payload(html, ("playersData",)):
        try:
            catalog = understat_catalog(html, season, identities)
            print(f"{label}: formato HTML legacy usato")
            return catalog
        except UpdateError:
            pass

    # 3) Formato corrente: JSON caricato dal frontend Understat.
    payload = download_understat_league_json(config, source_url, season, label)
    rows = [row for row in payload.get("players", []) if isinstance(row, dict)]
    return understat_catalog_from_rows(rows, season, identities)


def load_existing_enrichment(
    json_file: Path,
) -> tuple[dict[str, dict[str, dict[str, Any]]], dict[str, dict[str, Any]]]:
    advanced: dict[str, dict[str, dict[str, Any]]] = {}
    availability: dict[str, dict[str, Any]] = {}
    if not json_file.exists():
        return advanced, availability
    try:
        payload = json.loads(json_file.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return advanced, availability
    if not isinstance(payload, dict):
        return advanced, availability
    for role in ROLE_SHEETS:
        for player in payload.get(role, []) if isinstance(payload.get(role), list) else []:
            key = _player_id_key(player.get("id"))
            if not key:
                continue
            value = player.get("advanced")
            if isinstance(value, dict):
                advanced[key] = copy.deepcopy(value)
            value = player.get("availability")
            if isinstance(value, dict):
                availability[key] = copy.deepcopy(value)
    return advanced, availability


def download_advanced_catalogs(
    config: dict[str, Any],
    identities: list[dict[str, Any]],
    existing: dict[str, dict[str, dict[str, Any]]],
) -> tuple[dict[str, dict[str, dict[str, Any]]], list[str]]:
    """Carica metriche avanzate opzionali.

    Qualunque errore della fonte esterna (Understat, anti-bot, rete, parsing,
    Playwright/Chromium assente) produce solo un warning. I dati avanzati gia'
    presenti vengono mantenuti e l'aggiornamento principale prosegue.
    """
    sources = config.get("advanced_sources") or []
    merged = copy.deepcopy(existing)
    loaded: list[str] = []

    if not isinstance(sources, list):
        print(
            "ATTENZIONE: 'advanced_sources' non e' una lista. "
            "Salto i dati avanzati e continuo.",
            file=sys.stderr,
        )
        return merged, loaded

    for index, item in enumerate(sources, start=1):
        if not isinstance(item, dict) or not item.get("season") or not item.get("url"):
            print(
                f"ATTENZIONE: advanced_source #{index} non valida. "
                "Salto la fonte e continuo.",
                file=sys.stderr,
            )
            continue

        provider = str(item.get("provider", "understat")).casefold()
        season = str(item["season"]).strip()
        source_url = str(item["url"])

        try:
            if provider != "understat":
                raise UpdateError(f"Provider avanzato non supportato: {provider}")

            catalog = understat_catalog_current(
                config,
                source_url,
                season,
                identities,
                f"{provider} {season}",
            )

            if not catalog:
                raise UpdateError(f"Understat {season}: nessun giocatore collegato")

        except Exception as exc:
            # Fonte volutamente opzionale: non deve mai interrompere il workflow.
            print(
                f"ATTENZIONE: dati avanzati {provider} {season} non aggiornati: {exc}. "
                "Mantengo eventuali dati precedenti e continuo.",
                file=sys.stderr,
            )
            continue

        for player_id, details in catalog.items():
            merged.setdefault(player_id, {})[season] = details
        loaded.append(season)
        print(f"Understat {season}: {len(catalog)} giocatori collegati")

    return merged, loaded


STATUS_HEADINGS = {
    "infortunati": "Infortunato",
    "in dubbio": "In dubbio",
    "squalificati": "Squalificato",
    "diffidati": "Diffidato",
}
STATUS_BOUNDARIES = {
    "presentazione squadre",
    "dettaglio calciatori",
    "ballottaggi",
}


def parse_fantacalcio_availability(
    html: str,
    identities: list[dict[str, Any]],
    source_url: str,
    max_age_days: int = 21,
) -> tuple[dict[str, dict[str, Any]], str | None]:
    soup = BeautifulSoup(html, "html.parser")
    strings = [" ".join(x.split()) for x in soup.stripped_strings if x and x.strip()]
    joined = "\n".join(strings)
    dates = re.findall(r"Ultimo aggiornamento\s+(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}:\d{2})", joined, re.I)
    latest_dt: datetime | None = None
    for day, clock in dates:
        try:
            current = datetime.strptime(f"{day} {clock}", "%d/%m/%Y %H:%M")
        except ValueError:
            continue
        if latest_dt is None or current > latest_dt:
            latest_dt = current
    if latest_dt and (datetime.now() - latest_dt).days > max_age_days:
        raise UpdateError(
            f"Probabili formazioni Fantacalcio: dati vecchi ({latest_dt:%d/%m/%Y}), "
            f"oltre {max_age_days} giorni."
        )

    # Dizionario dei token esatti piu' frequenti nella pagina Fantacalcio.
    token_to_id: dict[str, str | None] = {}
    for token in strings:
        norm = " ".join(_name_tokens(token))
        if not norm or norm in token_to_id:
            continue
        matched = match_player_id(token, identities)
        token_to_id[norm] = matched

    result: dict[str, dict[str, Any]] = {}
    percent_re = re.compile(r"^(100|[1-9]?\d)%$")
    for i, token in enumerate(strings):
        player_id = token_to_id.get(" ".join(_name_tokens(token)))
        if not player_id:
            continue
        percentage = None
        for j in (i + 1, i - 1, i + 2, i - 2):
            if 0 <= j < len(strings):
                match = percent_re.match(strings[j])
                if match:
                    percentage = int(match.group(1))
                    break
        if percentage is not None:
            result.setdefault(player_id, {})["titolarita"] = percentage

    current_status: str | None = None
    for token in strings:
        norm = _norm_text(token)
        if norm in STATUS_HEADINGS:
            current_status = STATUS_HEADINGS[norm]
            continue
        if norm in STATUS_BOUNDARIES or norm.startswith("ultimo aggiornamento"):
            current_status = None
            continue
        if not current_status or norm == "nessun calciatore":
            continue
        player_id = token_to_id.get(" ".join(_name_tokens(token)))
        if player_id:
            result.setdefault(player_id, {})["stato"] = current_status

    source_updated = latest_dt.isoformat(timespec="minutes") if latest_dt else None
    for details in result.values():
        details["provider"] = "Fantacalcio.it"
        details["source"] = source_url
        if source_updated:
            details["updatedAt"] = source_updated
    if not result:
        raise UpdateError("Probabili formazioni Fantacalcio: nessun giocatore collegato.")
    return result, source_updated



def load_existing_set_pieces(json_file: Path) -> dict[str, dict[str, dict[str, Any]]]:
    result: dict[str, dict[str, dict[str, Any]]] = {}
    if not json_file.exists():
        return result
    try:
        payload = json.loads(json_file.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return result
    if not isinstance(payload, dict):
        return result
    for role in ROLE_SHEETS:
        for player in payload.get(role, []) if isinstance(payload.get(role), list) else []:
            key = _player_id_key(player.get("id"))
            value = player.get("setPieces")
            if key and isinstance(value, dict):
                result[key] = copy.deepcopy(value)
    return result


def _split_player_list(value: str) -> list[str]:
    value = re.sub(r"^[^:]+:\s*", "", value).strip().strip(". ")
    return [x.strip().strip(". ") for x in re.split(r"[,;]", value) if x.strip().strip(". ")]


def parse_fantacalcio_set_pieces(
    html: str,
    season: str,
    identities: list[dict[str, Any]],
    source_url: str,
) -> dict[str, dict[str, Any]]:
    """Legge gerarchie 'Rigoristi' e 'Calci da fermo/Punizioni/Corner' da una guida Fantacalcio."""
    soup = BeautifulSoup(html, "html.parser")
    result: dict[str, dict[str, Any]] = {}
    labels = {
        "rigoristi": "rigorista",
        "calci da fermo": "calciDaFermo",
        "punizioni": "punizioni",
        "corner": "corner",
    }

    # Gli articoli Fantacalcio espongono normalmente queste voci come testo. Analizziamo
    # ciascun nodo testuale, senza dipendere da classi CSS che possono cambiare.
    strings = [" ".join(x.split()) for x in soup.stripped_strings if x and x.strip()]
    known_teams = {_team_key(p["team"]): p["team"] for p in identities if p.get("team")}
    current_team: str | None = None
    for i, token in enumerate(strings):
        token_team = known_teams.get(_team_key(token))
        if token_team and len(_name_tokens(token)) <= 4:
            current_team = token_team
            continue

        low = token.casefold()
        label_key = next((key for key in labels if low.startswith(key + ":") or low == key), None)
        if not label_key:
            continue
        raw = token.split(":", 1)[1].strip() if ":" in token else ""
        if not raw and i + 1 < len(strings):
            raw = strings[i + 1]
        names = _split_player_list(raw)
        for rank, name in enumerate(names, start=1):
            player_id = match_player_id(name, identities, current_team) if current_team else None
            if not player_id:
                continue
            details = result.setdefault(player_id, {"provider": "Fantacalcio.it", "source": source_url, "season": season})
            details[f"{labels[label_key]}Rank"] = rank
    if not result:
        raise UpdateError(f"Calci piazzati Fantacalcio {season}: nessun giocatore collegato.")
    return result


def download_set_pieces_catalogs(
    config: dict[str, Any],
    identities: list[dict[str, Any]],
    existing: dict[str, dict[str, dict[str, Any]]],
) -> tuple[dict[str, dict[str, dict[str, Any]]], list[str]]:
    sources = config.get("set_pieces_sources") or []
    if not isinstance(sources, list):
        raise UpdateError("'set_pieces_sources' deve essere una lista.")
    merged = copy.deepcopy(existing)
    loaded: list[str] = []
    for item in sources:
        if not isinstance(item, dict) or not item.get("season") or not item.get("url"):
            raise UpdateError("Ogni set_pieces_source richiede 'season' e 'url'.")
        season = str(item["season"]).strip()
        required = bool(item.get("required", False))
        use_cookie = bool(item.get("use_cookie", False))
        url = str(item["url"])
        try:
            html = download_text(
                config,
                url,
                f"Calci piazzati Fantacalcio {season}",
                use_cookie=use_cookie,
            )
            catalog = parse_fantacalcio_set_pieces(html, season, identities, url)
        except (UpdateError, requests.RequestException) as exc:
            if required:
                raise
            print(f"ATTENZIONE: {exc}. Mantengo eventuali calci piazzati {season}.", file=sys.stderr)
            continue
        for player_id, details in catalog.items():
            merged.setdefault(player_id, {})[season] = details
        loaded.append(season)
        print(f"Calci piazzati {season}: {len(catalog)} giocatori collegati")
    return merged, loaded


def download_availability_catalog(
    config: dict[str, Any],
    identities: list[dict[str, Any]],
    existing: dict[str, dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], str | None, bool]:
    existing_updated = max(
        (str(v.get("updatedAt")) for v in existing.values() if v.get("updatedAt")),
        default=None,
    )
    item = config.get("availability_source")
    if not item:
        return copy.deepcopy(existing), existing_updated, False
    if not isinstance(item, dict) or not item.get("url"):
        raise UpdateError("'availability_source' richiede almeno 'url'.")
    required = bool(item.get("required", False))
    url = str(item["url"])
    try:
        html = download_text(
            config,
            url,
            "Probabili formazioni Fantacalcio",
            use_cookie=bool(item.get("use_cookie", False)),
        )
        parsed, updated_at = parse_fantacalcio_availability(
            html,
            identities,
            url,
            max_age_days=int(item.get("max_age_days", 21)),
        )
        print(f"Probabili formazioni: {len(parsed)} giocatori collegati")
        return parsed, updated_at, True
    except (UpdateError, requests.RequestException) as exc:
        if required:
            raise
        print(f"ATTENZIONE: {exc}. Mantengo eventuali dati disponibilita'.", file=sys.stderr)
        return copy.deepcopy(existing), existing_updated, False



def load_existing_injuries(json_file: Path) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    if not json_file.exists():
        return result
    try:
        payload = json.loads(json_file.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return result
    if not isinstance(payload, dict):
        return result
    for role in ROLE_SHEETS:
        for player in payload.get(role, []) if isinstance(payload.get(role), list) else []:
            key = _player_id_key(player.get("id"))
            value = player.get("injury")
            if key and isinstance(value, dict):
                result[key] = copy.deepcopy(value)
    return result


def _injury_return_text(description: str) -> str:
    text = " ".join(str(description or "").split()).strip()
    if not text:
        return ""
    sentences = [x.strip() for x in re.split(r"(?<=[.!?])\s+", text) if x.strip()]
    keywords = (
        "rientro", "rientrare", "tornare", "recuperabile", "convocabile",
        "arruolabile", "tempi di recupero", "stagione finita", "campionato finito",
        "da valutare", "out ", "fuori causa", "non prima", "proverà a recuperare",
        "provera a recuperare", "punta a tornare", "può tornare", "puo tornare",
    )
    for sentence in sentences:
        low = sentence.casefold()
        if any(key in low for key in keywords):
            return sentence
    return ""


def parse_fantacalcio_injuries(
    html: str,
    identities: list[dict[str, Any]],
    source_url: str,
) -> tuple[dict[str, dict[str, Any]], str]:
    """Legge la pagina ufficiale Fantacalcio degli infortunati.

    Il parser usa il testo della sezione principale, non classi CSS: e' meno fragile
    rispetto a cambi di layout. Il matching resta conservativo per nome + squadra.
    """
    soup = BeautifulSoup(html, "html.parser")
    strings = [" ".join(x.split()) for x in soup.stripped_strings if x and x.strip()]
    if not strings:
        raise UpdateError("Infortunati Fantacalcio: pagina vuota.")

    start = next((i for i, token in enumerate(strings) if _norm_text(token) == "infortunati serie a"), None)
    if start is None:
        raise UpdateError("Infortunati Fantacalcio: intestazione principale non trovata.")

    end = len(strings)
    for i in range(start + 1, len(strings)):
        norm = _norm_text(strings[i])
        if norm.startswith("i calciatori infortunati sono") or norm == "prossimo turno":
            end = i
            break

    known_teams = {_team_key(p.get("team")): str(p.get("team") or "") for p in identities if p.get("team")}
    fetched_at = datetime.now().astimezone().isoformat(timespec="minutes")

    # Una lettura riuscita della pagina consente di distinguere anche chi non e' elencato.
    result: dict[str, dict[str, Any]] = {}
    for identity in identities:
        player_id = _player_id_key(identity.get("id"))
        if not player_id:
            continue
        result[player_id] = {
            "injured": False,
            "status": "Disponibile",
            "provider": "Fantacalcio.it",
            "source": source_url,
            "updatedAt": fetched_at,
        }

    current_team: str | None = None
    matched = 0
    i = start + 1
    while i < end:
        token = strings[i]
        norm = _norm_text(token)
        team = known_teams.get(_team_key(token))
        if team:
            current_team = team
            i += 1
            continue
        if not current_team or norm in {"nessuno", "* * *"}:
            i += 1
            continue

        player_id = match_player_id(token, identities, current_team)
        if not player_id:
            i += 1
            continue

        description = ""
        if i + 1 < end:
            candidate = strings[i + 1]
            candidate_norm = _norm_text(candidate)
            candidate_team = known_teams.get(_team_key(candidate))
            candidate_player = match_player_id(candidate, identities, current_team) if not candidate_team else None
            if not candidate_team and not candidate_player and candidate_norm not in {"nessuno", "* * *"}:
                description = candidate

        details = {
            "injured": True,
            "status": "Infortunato",
            "description": description,
            "returnText": _injury_return_text(description),
            "provider": "Fantacalcio.it",
            "source": source_url,
            "updatedAt": fetched_at,
        }
        result[player_id] = details
        matched += 1
        i += 2 if description else 1

    if matched == 0:
        # In piena stagione una pagina valida ma senza match e' sospetta: non azzerare dati precedenti.
        raise UpdateError("Infortunati Fantacalcio: nessun giocatore collegato alla rosa corrente.")
    return result, fetched_at


def download_injury_catalog(
    config: dict[str, Any],
    identities: list[dict[str, Any]],
    existing: dict[str, dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], str | None, bool]:
    existing_updated = max(
        (str(v.get("updatedAt")) for v in existing.values() if v.get("updatedAt")),
        default=None,
    )
    item = config.get("injury_source")
    if not item:
        return copy.deepcopy(existing), existing_updated, False
    if not isinstance(item, dict) or not item.get("url"):
        print("ATTENZIONE: 'injury_source' non valida. Mantengo eventuali dati infortuni.", file=sys.stderr)
        return copy.deepcopy(existing), existing_updated, False

    url = str(item["url"])
    required = bool(item.get("required", False))
    try:
        html = download_text(
            config,
            url,
            "Infortunati Fantacalcio",
            use_cookie=bool(item.get("use_cookie", False)),
        )
        parsed, updated_at = parse_fantacalcio_injuries(html, identities, url)
        injured_count = sum(1 for value in parsed.values() if value.get("injured") is True)
        print(f"Infortunati Fantacalcio: {injured_count} giocatori segnalati")
        return parsed, updated_at, True
    except (UpdateError, requests.RequestException) as exc:
        if required:
            raise
        print(f"ATTENZIONE: {exc}. Mantengo eventuali dati infortuni e continuo.", file=sys.stderr)
        return copy.deepcopy(existing), existing_updated, False


def build_player_catalog(
    workbook_file: Path,
    stats_by_id: dict[str, dict[str, dict[str, Any]]] | None = None,
    advanced_by_id: dict[str, dict[str, dict[str, Any]]] | None = None,
    availability_by_id: dict[str, dict[str, Any]] | None = None,
    set_pieces_by_id: dict[str, dict[str, dict[str, Any]]] | None = None,
    availability_updated_at: str | None = None,
    injury_by_id: dict[str, dict[str, Any]] | None = None,
    injury_updated_at: str | None = None,
) -> dict[str, Any]:
    """Legge i quattro fogli ruolo e genera il catalogo pubblico dei giocatori."""
    try:
        wb = load_workbook(workbook_file, read_only=False, data_only=True)
    except Exception as exc:
        raise UpdateError(f"Impossibile aprire il workbook per generare players.json: {exc}") from exc

    stats_by_id = stats_by_id or {}
    advanced_by_id = advanced_by_id or {}
    availability_by_id = availability_by_id or {}
    set_pieces_by_id = set_pieces_by_id or {}
    injury_by_id = injury_by_id or {}
    lookup = sheet_lookup(wb)
    catalog: dict[str, Any] = {}
    seasons: set[str] = set()
    advanced_seasons: set[str] = set()
    set_pieces_seasons: set[str] = set()
    try:
        for role, expected_sheet in ROLE_SHEETS.items():
            real_name = lookup.get(expected_sheet.casefold())
            if real_name is None:
                raise UpdateError(
                    f"Foglio '{expected_sheet}' mancante: impossibile generare players.json."
                )

            ws = wb[real_name]
            headers = {
                str(ws.cell(HEADER_ROW, col).value): col
                for col in range(DATA_FIRST_COL, DATA_LAST_COL + 1)
                if ws.cell(HEADER_ROW, col).value is not None
            }
            missing = [name for name in ("Id", "Nome", "Squadra", "FVM") if name not in headers]
            if missing:
                raise UpdateError(
                    f"Colonne mancanti nel foglio '{real_name}' per players.json: {', '.join(missing)}"
                )

            players: list[dict[str, Any]] = []
            for row in range(HEADER_ROW + 1, ws.max_row + 1):
                player_id = ws.cell(row, headers["Id"]).value
                name = ws.cell(row, headers["Nome"]).value
                if player_id is None or not name:
                    continue

                info = {
                    header: _json_value(ws.cell(row, col).value)
                    for header, col in headers.items()
                }
                key = _player_id_key(player_id)
                player_stats = copy.deepcopy(stats_by_id.get(key, {}))
                player_advanced = copy.deepcopy(advanced_by_id.get(key, {}))
                player_availability = copy.deepcopy(availability_by_id.get(key, {}))
                player_set_pieces = copy.deepcopy(set_pieces_by_id.get(key, {}))
                player_injury = copy.deepcopy(injury_by_id.get(key, {}))
                seasons.update(player_stats.keys())
                advanced_seasons.update(player_advanced.keys())
                set_pieces_seasons.update(player_set_pieces.keys())
                players.append(
                    {
                        "id": _json_number(player_id),
                        "name": str(name).strip(),
                        "team": str(ws.cell(row, headers["Squadra"] ).value or "").strip(),
                        "fvm": _json_number(ws.cell(row, headers["FVM"]).value),
                        "info": info,
                        "stats": player_stats,
                        "advanced": player_advanced,
                        "availability": player_availability,
                        "setPieces": player_set_pieces,
                        "injury": player_injury,
                    }
                )
            catalog[role] = players
    finally:
        wb.close()

    catalog["_meta"] = {
        "statsSeasons": sorted(seasons, reverse=True),
        "advancedSeasons": sorted(advanced_seasons, reverse=True),
        "setPiecesSeasons": sorted(set_pieces_seasons, reverse=True),
        "availabilityUpdatedAt": availability_updated_at,
        "injuryUpdatedAt": injury_updated_at,
    }
    return catalog


def update_players_json(
    workbook_file: Path,
    json_file: Path,
    stats_by_id: dict[str, dict[str, dict[str, Any]]] | None = None,
    advanced_by_id: dict[str, dict[str, dict[str, Any]]] | None = None,
    availability_by_id: dict[str, dict[str, Any]] | None = None,
    set_pieces_by_id: dict[str, dict[str, dict[str, Any]]] | None = None,
    availability_updated_at: str | None = None,
    injury_by_id: dict[str, dict[str, Any]] | None = None,
    injury_updated_at: str | None = None,
) -> tuple[bool, dict[str, int]]:
    """Genera players.json in modo atomico e lo riscrive solo quando i dati cambiano."""
    catalog = build_player_catalog(
        workbook_file,
        stats_by_id,
        advanced_by_id,
        availability_by_id,
        set_pieces_by_id,
        availability_updated_at,
        injury_by_id,
        injury_updated_at,
    )
    current_text = json_file.read_text(encoding="utf-8") if json_file.exists() else None
    current_catalog = None
    if current_text:
        try:
            current_catalog = json.loads(current_text)
        except json.JSONDecodeError:
            current_catalog = None

    # catalogUpdatedAt descrive l'ultimo cambiamento effettivo del catalogo.
    # Viene escluso dal confronto per evitare commit inutili a ogni esecuzione.
    current_for_compare = copy.deepcopy(current_catalog) if isinstance(current_catalog, dict) else None
    if isinstance(current_for_compare, dict):
        current_for_compare.setdefault("_meta", {}).pop("catalogUpdatedAt", None)
    catalog_for_compare = copy.deepcopy(catalog)
    catalog_for_compare.setdefault("_meta", {}).pop("catalogUpdatedAt", None)
    has_catalog_timestamp = bool(
        isinstance(current_catalog, dict)
        and isinstance(current_catalog.get("_meta"), dict)
        and current_catalog["_meta"].get("catalogUpdatedAt")
    )
    changed = current_for_compare != catalog_for_compare or not has_catalog_timestamp

    if changed:
        catalog.setdefault("_meta", {})["catalogUpdatedAt"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
        payload = json.dumps(catalog, ensure_ascii=False, indent=2) + "\n"
        json_file.parent.mkdir(parents=True, exist_ok=True)
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{json_file.stem}_",
            suffix=json_file.suffix or ".json",
            dir=str(json_file.parent),
        )
        os.close(fd)
        temp_path = Path(temp_name)
        try:
            temp_path.write_text(payload, encoding="utf-8")
            os.replace(temp_path, json_file)
        finally:
            if temp_path.exists():
                temp_path.unlink()

    return changed, {role: len(catalog[role]) for role in ROLE_SHEETS}

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scarica il listone Fantacalcio e aggiorna players.json senza persistere file XLSX."
    )
    parser.add_argument(
        "--config",
        default="config.json",
        help="Percorso del file JSON di configurazione (default: config.json)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    config_path = Path(args.config).expanduser().resolve()

    try:
        config = load_config(config_path)
        config_dir = config_path.parent
        players_file = resolve_path(str(config["players_file"]), config_dir)

        source_bytes = download_source(config)

        # Il listone XLSX e' solo una sorgente temporanea: non viene mai
        # scritto nel repository o mantenuto come artefatto del progetto.
        fd, temp_name = tempfile.mkstemp(prefix="fantacalcio_prices_", suffix=".xlsx")
        os.close(fd)
        source_file = Path(temp_name)
        try:
            source_file.write_bytes(source_bytes)

            # Validazione esplicita della struttura sorgente prima di generare il JSON.
            try:
                source_wb = load_workbook(source_file, read_only=False, data_only=False)
                resolved = validate_source_workbook(source_wb)
                source_counts = {
                    name: count_data_rows(source_wb[real_name])
                    for name, real_name in resolved.items()
                }
                source_wb.close()
            except Exception as exc:
                raise UpdateError(f"Impossibile validare il listone sorgente: {exc}") from exc

            identities = load_player_identities(source_file)

            existing_stats = load_existing_stats(players_file)
            stats_by_id, loaded_stats_seasons = download_stats_catalogs(config, existing_stats)

            existing_advanced, existing_availability = load_existing_enrichment(players_file)
            advanced_by_id, loaded_advanced_seasons = download_advanced_catalogs(
                config, identities, existing_advanced
            )
            availability_by_id, availability_updated_at, availability_refreshed = (
                download_availability_catalog(config, identities, existing_availability)
            )

            existing_injuries = load_existing_injuries(players_file)
            injury_by_id, injury_updated_at, injury_refreshed = download_injury_catalog(
                config, identities, existing_injuries
            )

            existing_set_pieces = load_existing_set_pieces(players_file)
            set_pieces_by_id, loaded_set_pieces_seasons = download_set_pieces_catalogs(
                config, identities, existing_set_pieces
            )

            players_changed, players_stats = update_players_json(
                source_file,
                players_file,
                stats_by_id,
                advanced_by_id,
                availability_by_id,
                set_pieces_by_id,
                availability_updated_at,
                injury_by_id,
                injury_updated_at,
            )
        finally:
            try:
                source_file.unlink(missing_ok=True)
            except OSError:
                pass

        status = "aggiornato" if players_changed else "gia' allineato"
        print(f"Catalogo JSON {status}: {players_file}")
        for role in ("P", "D", "C", "A"):
            print(f"  - {role}: {players_stats[role]} giocatori disponibili")
        if loaded_stats_seasons:
            print("  - statistiche Fantacalcio aggiornate: " + ", ".join(loaded_stats_seasons))
        if loaded_advanced_seasons:
            print("  - metriche Understat aggiornate: " + ", ".join(loaded_advanced_seasons))
        if availability_refreshed:
            print("  - disponibilita' / titolarita' Fantacalcio aggiornate")
        if injury_refreshed:
            print("  - infortuni Fantacalcio aggiornati")
        if loaded_set_pieces_seasons:
            print("  - gerarchie calci piazzati aggiornate: " + ", ".join(loaded_set_pieces_seasons))
        for sheet_name in SHEETS_TO_UPDATE:
            print(f"- sorgente {sheet_name}: {source_counts[sheet_name]} giocatori")
        return 0

    except (UpdateError, requests.RequestException, OSError) as exc:
        print(f"ERRORE: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
